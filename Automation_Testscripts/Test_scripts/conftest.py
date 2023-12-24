import configparser

import pytest
from openpyxl import load_workbook
from selenium.webdriver.chrome import webdriver
from selenium.webdriver.chrome.service import Service
from selenium import webdriver


@pytest.fixture(scope='class')
def setup(request):
    service_obj = Service("/home/ezetape/Downloads/chromedriver-linux64_backup/chromedriver")
    driver = webdriver.Chrome(service=service_obj)
    driver.implicitly_wait(5)
    config_parse = configparser.ConfigParser()
    config_parse.read('config.ini')
    envirurl = config_parse.get('URls', 'URL')
    print(envirurl)
    driver.get(envirurl)
    driver.maximize_window()
    request.cls.driver = driver

    yield
    driver.close()


@pytest.fixture()
def read_test_data(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_data.append(row)

    return test_data
